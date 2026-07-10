#!/usr/bin/env python3
"""
Generate site product data from the Loop Excel catalog.

Only rows with local downloaded Product Image paths are exported. Rows that
still contain external search formulas are intentionally skipped so the site
does not render broken or placeholder product cards.
"""

from __future__ import annotations

import argparse
import json
import re
from pathlib import Path
from typing import Any

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_WORKBOOK = ROOT / "loop_300_product_catalog.xlsx"
DEFAULT_OUTPUT = ROOT / "catalog-products.js"
SHEET_NAME = "Product Catalog"
REQUIRED_COLUMNS = [
    "Product Name",
    "Brand",
    "Main Category",
    "Subcategory",
    "Size / Unit",
    "Price",
    "Product Image",
]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Generate catalog-products.js from the Loop workbook.")
    parser.add_argument("--workbook", type=Path, default=DEFAULT_WORKBOOK)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    return parser.parse_args()


def clean_text(value: Any) -> str:
    return str(value or "").strip()


def slugify(value: str) -> str:
    value = value.lower()
    value = re.sub(r"[^a-z0-9]+", "-", value)
    value = re.sub(r"-{2,}", "-", value).strip("-")
    return value or "product"


def is_local_image(value: str) -> bool:
    normalized = value.replace("\\", "/").strip()
    return normalized.startswith("images/") and normalized.lower().endswith((".jpg", ".jpeg", ".png", ".webp"))


def get_headers(sheet: Any) -> dict[str, int]:
    headers = {clean_text(cell.value): cell.column for cell in sheet[1] if clean_text(cell.value)}
    missing = [column for column in REQUIRED_COLUMNS if column not in headers]
    if missing:
        raise RuntimeError(f"Missing required columns: {', '.join(missing)}")
    return headers


def product_from_row(sheet: Any, row_index: int, headers: dict[str, int]) -> dict[str, Any] | None:
    row = {
        column: sheet.cell(row=row_index, column=headers[column]).value
        for column in REQUIRED_COLUMNS
    }
    image = clean_text(row["Product Image"]).replace("\\", "/")
    if not is_local_image(image):
        return None

    name = clean_text(row["Product Name"])
    unit = clean_text(row["Size / Unit"])
    return {
        "id": slugify(f"{name} {unit}"),
        "name": name,
        "brand": clean_text(row["Brand"]),
        "category": clean_text(row["Main Category"]),
        "subcategory": clean_text(row["Subcategory"]),
        "unit": unit,
        "price": int(float(row["Price"] or 0)),
        "image": image,
    }


def main() -> int:
    args = parse_args()
    workbook = openpyxl.load_workbook(args.workbook, data_only=False)
    sheet = workbook[SHEET_NAME]
    headers = get_headers(sheet)

    products: list[dict[str, Any]] = []
    seen_ids: set[str] = set()
    for row_index in range(2, sheet.max_row + 1):
        product = product_from_row(sheet, row_index, headers)
        if product is None:
            continue
        if product["id"] in seen_ids:
            raise RuntimeError(f"Duplicate generated product id: {product['id']}")
        seen_ids.add(product["id"])
        products.append(product)

    payload = json.dumps(products, ensure_ascii=False, indent=2)
    args.output.write_text(
        "window.LOOP_CATALOG_PRODUCTS = "
        + payload
        + ";\n",
        encoding="utf-8",
    )
    print(f"Generated {len(products)} products at {args.output}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
