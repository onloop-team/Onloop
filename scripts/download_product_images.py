#!/usr/bin/env python3
"""
Download real product images for Loop catalog rows using SerpAPI.

The script reads loop_300_product_catalog.xlsx, downloads matching product
images into images/catalog-products, and updates the Product Image column with
local paths. It creates a timestamped workbook backup before saving changes.
"""

from __future__ import annotations

import argparse
import csv
import mimetypes
import os
import re
import shutil
import sys
import time
from datetime import datetime
from pathlib import Path
from typing import Any
from urllib.parse import urlparse

import openpyxl
import requests


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_WORKBOOK = ROOT / "loop_300_product_catalog.xlsx"
DEFAULT_OUTPUT_DIR = ROOT / "images" / "catalog-products"
DEFAULT_FAILURE_LOG = ROOT / "image-download-failures.csv"
SHEET_NAME = "Product Catalog"
REQUIRED_COLUMNS = [
    "Product Name",
    "Brand",
    "Main Category",
    "Subcategory",
    "Size / Unit",
    "Product Image",
]
SERPAPI_ENDPOINT = "https://serpapi.com/search.json"
REQUEST_TIMEOUT = 25
IMAGE_TIMEOUT = 30
VALID_CONTENT_TYPES = {
    "image/jpeg": ".jpg",
    "image/jpg": ".jpg",
    "image/png": ".png",
    "image/webp": ".webp",
}
USER_AGENT = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
    "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0 Safari/537.36"
)


class RateLimitError(RuntimeError):
    pass


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Search SerpAPI and download product images for the Loop catalog."
    )
    parser.add_argument("--workbook", type=Path, default=DEFAULT_WORKBOOK)
    parser.add_argument(
        "--output-workbook",
        type=Path,
        default=None,
        help="Save updates to this workbook instead of overwriting --workbook.",
    )
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT_DIR)
    parser.add_argument("--failure-log", type=Path, default=DEFAULT_FAILURE_LOG)
    parser.add_argument("--limit", type=int, default=None, help="Process at most N rows.")
    parser.add_argument("--delay", type=float, default=0.8, help="Delay between rows in seconds.")
    parser.add_argument("--save-every", type=int, default=25, help="Save workbook progress every N updates.")
    parser.add_argument("--overwrite", action="store_true", help="Replace existing Product Image values.")
    parser.add_argument(
        "--offline-existing-only",
        action="store_true",
        help="Only map already downloaded files into the workbook; do not call SerpAPI.",
    )
    parser.add_argument("--dry-run", action="store_true", help="Preview rows and queries without downloading or saving.")
    return parser.parse_args()


def clean_text(value: Any) -> str:
    return str(value or "").strip()


def slugify(parts: list[str]) -> str:
    text = " ".join(part for part in parts if part).lower()
    text = re.sub(r"[^a-z0-9]+", "-", text)
    text = re.sub(r"-{2,}", "-", text).strip("-")
    return text[:90] or "product"


def product_file_base(row: dict[str, str]) -> str:
    brand = row["Brand"]
    product_name = row["Product Name"]
    if brand and product_name.lower().startswith(brand.lower()):
        return slugify([product_name, row["Size / Unit"]])
    return slugify([brand, product_name, row["Size / Unit"]])


def build_query(row: dict[str, str]) -> str:
    brand = row["Brand"]
    product_name = row["Product Name"]
    if brand and product_name.lower().startswith(brand.lower()):
        terms = [product_name]
    else:
        terms = [brand, product_name]
    terms.extend([row["Size / Unit"], row["Subcategory"], "product image", "Nigeria"])
    return " ".join(term for term in terms if term)


def get_headers(sheet: Any) -> dict[str, int]:
    headers: dict[str, int] = {}
    for cell in sheet[1]:
        label = clean_text(cell.value)
        if label:
            headers[label] = cell.column
    missing = [name for name in REQUIRED_COLUMNS if name not in headers]
    if missing:
        raise RuntimeError(f"Missing required columns: {', '.join(missing)}")
    return headers


def row_to_product(sheet: Any, row_index: int, headers: dict[str, int]) -> dict[str, str]:
    return {
        column: clean_text(sheet.cell(row=row_index, column=headers[column]).value)
        for column in REQUIRED_COLUMNS
    }


def serpapi_search(api_key: str, query: str) -> list[dict[str, Any]]:
    params = {
        "engine": "google_images",
        "q": query,
        "api_key": api_key,
        "ijn": "0",
        "safe": "active",
    }
    response = requests.get(SERPAPI_ENDPOINT, params=params, timeout=REQUEST_TIMEOUT)
    if response.status_code == 429:
        raise RateLimitError("SerpAPI returned 429 Too Many Requests. Wait before resuming.")
    response.raise_for_status()
    data = response.json()
    if "error" in data:
        raise RuntimeError(data["error"])
    return data.get("images_results", []) or []


def candidate_urls(results: list[dict[str, Any]]) -> list[str]:
    urls: list[str] = []
    for result in results:
        for key in ("original", "thumbnail"):
            url = result.get(key)
            if isinstance(url, str) and url.startswith(("http://", "https://")):
                urls.append(url)
    return list(dict.fromkeys(urls))


def extension_from_response(url: str, content_type: str) -> str:
    base_type = content_type.split(";")[0].strip().lower()
    if base_type in VALID_CONTENT_TYPES:
        return VALID_CONTENT_TYPES[base_type]

    suffix = Path(urlparse(url).path).suffix.lower()
    if suffix in {".jpg", ".jpeg", ".png", ".webp"}:
        return ".jpg" if suffix == ".jpeg" else suffix

    guessed = mimetypes.guess_extension(base_type)
    if guessed in {".jpg", ".jpeg", ".png", ".webp"}:
        return ".jpg" if guessed == ".jpeg" else guessed
    return ".jpg"


def download_first_image(urls: list[str], output_dir: Path, base_name: str) -> tuple[Path, str]:
    headers = {"User-Agent": USER_AGENT}
    last_error = "No image URLs returned"

    for index, url in enumerate(urls[:8], start=1):
        try:
            response = requests.get(url, headers=headers, timeout=IMAGE_TIMEOUT, stream=True)
            response.raise_for_status()
            content_type = response.headers.get("Content-Type", "").split(";")[0].lower()
            if content_type and content_type not in VALID_CONTENT_TYPES:
                last_error = f"Unsupported content type {content_type} from {url}"
                continue

            content = response.content
            if len(content) < 1500:
                last_error = f"Image too small from {url}"
                continue

            extension = extension_from_response(url, content_type)
            filename = f"{base_name}{extension}" if index == 1 else f"{base_name}-{index}{extension}"
            path = output_dir / filename
            path.write_bytes(content)
            return path, url
        except requests.RequestException as exc:
            last_error = f"{type(exc).__name__}: {exc}"

    raise RuntimeError(last_error)


def find_existing_image(output_dir: Path, base_name: str) -> Path | None:
    for extension in (".jpg", ".png", ".webp", ".jpeg"):
        path = output_dir / f"{base_name}{extension}"
        if path.exists():
            return path
    matches = sorted(output_dir.glob(f"{base_name}-*"))
    for path in matches:
        if path.suffix.lower() in {".jpg", ".jpeg", ".png", ".webp"}:
            return path
    return None


def workbook_backup_path(workbook: Path) -> Path:
    timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    return workbook.with_name(f"{workbook.stem}.backup-{timestamp}{workbook.suffix}")


def relative_path(path: Path) -> str:
    return path.resolve().relative_to(ROOT.resolve()).as_posix()


def has_local_image_path(value: str) -> bool:
    if not value:
        return False

    lowered = value.lower().strip()
    if lowered.startswith("=hyperlink("):
        return False
    if lowered.startswith(("http://", "https://")):
        return False
    return lowered.endswith((".jpg", ".jpeg", ".png", ".webp")) or "images/" in lowered.replace("\\", "/")


def write_failures(path: Path, failures: list[dict[str, str]]) -> None:
    if not failures:
        if path.exists():
            path.unlink()
        return

    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=["row", "product_name", "brand", "query", "reason"],
        )
        writer.writeheader()
        writer.writerows(failures)


def main() -> int:
    args = parse_args()
    api_key = os.environ.get("SERPAPI_KEY", "").strip()

    if not args.workbook.exists():
        print(f"Workbook not found: {args.workbook}", file=sys.stderr)
        return 1
    output_workbook = args.output_workbook or args.workbook

    if not api_key and not args.dry_run and not args.offline_existing_only:
        print("SERPAPI_KEY environment variable is required.", file=sys.stderr)
        return 1

    workbook = openpyxl.load_workbook(args.workbook)
    if SHEET_NAME not in workbook.sheetnames:
        print(f"Sheet not found: {SHEET_NAME}", file=sys.stderr)
        return 1

    sheet = workbook[SHEET_NAME]
    headers = get_headers(sheet)
    image_col = headers["Product Image"]
    args.output_dir.mkdir(parents=True, exist_ok=True)

    processed = 0
    downloaded = 0
    skipped = 0
    reused = 0
    updates_since_save = 0
    backup: Path | None = None
    failures: list[dict[str, str]] = []

    def ensure_backup() -> None:
        nonlocal backup
        if backup is None and output_workbook.resolve() == args.workbook.resolve():
            backup = workbook_backup_path(args.workbook)
            shutil.copy2(args.workbook, backup)

    def save_progress(force: bool = False) -> None:
        nonlocal updates_since_save
        if updates_since_save == 0 and not force:
            return
        ensure_backup()
        output_workbook.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(output_workbook)
        updates_since_save = 0
        print(f"  progress saved to {output_workbook}", flush=True)

    for row_index in range(2, sheet.max_row + 1):
        if args.limit is not None and processed >= args.limit:
            break

        product = row_to_product(sheet, row_index, headers)
        existing_image = product["Product Image"]
        if has_local_image_path(existing_image) and not args.overwrite:
            skipped += 1
            continue

        product_name = product["Product Name"]
        if not product_name:
            continue

        processed += 1
        query = build_query(product)
        base_name = product_file_base(product)
        print(f"[{row_index}] {query}", flush=True)

        if args.dry_run:
            continue

        try:
            existing_file = find_existing_image(args.output_dir, base_name)
            if existing_file is not None:
                sheet.cell(row=row_index, column=image_col).value = relative_path(existing_file)
                reused += 1
                updates_since_save += 1
                print(f"  reused {relative_path(existing_file)}", flush=True)
                if args.save_every > 0 and updates_since_save >= args.save_every:
                    save_progress()
                continue

            if args.offline_existing_only:
                failures.append(
                    {
                        "row": str(row_index),
                        "product_name": product_name,
                        "brand": product["Brand"],
                        "query": query,
                        "reason": "No existing downloaded image file found; offline mode did not call SerpAPI.",
                    }
                )
                print("  missing existing image file; skipped SerpAPI in offline mode", flush=True)
                continue

            results = serpapi_search(api_key, query)
            path, source_url = download_first_image(candidate_urls(results), args.output_dir, base_name)
            sheet.cell(row=row_index, column=image_col).value = relative_path(path)
            downloaded += 1
            updates_since_save += 1
            print(f"  saved {relative_path(path)}", flush=True)
            print(f"  source {source_url}", flush=True)
            if args.save_every > 0 and updates_since_save >= args.save_every:
                save_progress()
        except RateLimitError as exc:
            failures.append(
                {
                    "row": str(row_index),
                    "product_name": product_name,
                    "brand": product["Brand"],
                    "query": query,
                    "reason": str(exc),
                }
            )
            print(f"  stopped: {exc}", flush=True)
            break
        except Exception as exc:
            failures.append(
                {
                    "row": str(row_index),
                    "product_name": product_name,
                    "brand": product["Brand"],
                    "query": query,
                    "reason": str(exc),
                }
            )
            print(f"  failed: {exc}", flush=True)

        if args.delay > 0:
            time.sleep(args.delay)

    if args.dry_run:
        print(f"Dry run complete. Rows previewed: {processed}. No files changed.")
        return 0

    if downloaded or reused or updates_since_save:
        save_progress(force=True)
        if backup is not None:
            print(f"Workbook backup: {backup}")
        print(f"Workbook updated: {output_workbook}")
    else:
        print("No workbook changes to save.")

    write_failures(args.failure_log, failures)
    if failures:
        print(f"Failures logged: {args.failure_log}")

    print(
        f"Done. Processed: {processed}. Downloaded: {downloaded}. "
        f"Reused existing files: {reused}. Skipped existing paths: {skipped}. Failed: {len(failures)}."
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
